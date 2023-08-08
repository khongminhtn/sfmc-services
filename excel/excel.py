from openpyxl import load_workbook


def read_excel(name):
    # Load work book
    wb = load_workbook(name)

    # Load worksheet
    ws_name = wb.sheetnames[0]
    ws = wb[ws_name]

    # Get data from worksheet
    data = []
    for row in ws.iter_cols(values_only=True):
        data.append({
            "header": row[0],
            "values": row[1:]
    })

    return data