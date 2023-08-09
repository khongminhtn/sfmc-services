from openpyxl import load_workbook


def read_excel(location):
    '''
    Ingest translations in excel format.

        Param:
            location (string): location of excel file xlsm
        
        Returns:
            translations(dictionary): carry info extracted from excel
    '''
    # Load work book
    wb = load_workbook(location)

    # Load worksheet
    ws_name = wb.sheetnames[0]
    ws = wb[ws_name]

    # Get data from worksheet
    translations = []
    for row in ws.iter_cols(values_only=True):
        translations.append({
            "header": row[0],
            "values": row[1:]
    })

    return translations